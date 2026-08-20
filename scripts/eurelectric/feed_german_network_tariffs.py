# SPDX-FileCopyrightText: Contributors to PyPSA-Eurelectric
#
# SPDX-License-Identifier: MIT
"""
Feed German PyPSA grid-study *physical* results into a copy of
german_network_tariff_model_v3.xlsx.

This script does not modify the original workbook.

The Excel remains the absolute German financial/regulatory baseline
(RAB, WACC, depreciation, OPEX, Inv_base). PyPSA is used only for
endogenous physical ratios (German DSO capacity/peak flow, TSO volume,
end-use electricity). PyPSA absolute grid CAPEX is never written in.

Usage (repo root, pixi env):

    python scripts/eurelectric/feed_german_network_tariffs.py
"""

from __future__ import annotations

import argparse
import logging
import math
import shutil
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import pypsa

logger = logging.getLogger(__name__)

RUN_META: dict[str, dict] = {
    "grid-build-2040": {"label": "Build", "cap_fraction": 1.00},
    "grid-inaction-2040-mild85": {"label": "85%", "cap_fraction": 0.85},
    "grid-inaction-2040-mild75": {"label": "75%", "cap_fraction": 0.75},
    "grid-inaction-2040-mild65": {"label": "65%", "cap_fraction": 0.65},
    "grid-inaction-2040-moderate": {"label": "50%", "cap_fraction": 0.50},
}

# Europe-wide annualised *new* DSO investment cap (bn EUR/yr). The cap is not
# Germany-specific. Used only to record whether the EU constraint binds.
EU_DSO_BUDGET_BN = {
    ("85%", 2030): 26.066,
    ("85%", 2040): 12.056,
    ("75%", 2030): 23.000,
    ("75%", 2040): 10.638,
    ("65%", 2030): 19.933,
    ("65%", 2040): 9.220,
    ("50%", 2030): 15.333,
    ("50%", 2040): 7.092,
}

# Documented IMK peak-year split (Parameters C21): ÜNB ~19.8 + VNB ~14.4.
# Normalised so the two shares sum to 1.
TSO_SHARE = 19.8 / 34.2
DSO_SHARE = 14.4 / 34.2

HISTORICAL_THROUGH = 2025
PYPSA_2030_WINDOW = range(2026, 2031)
PYPSA_2040_WINDOW = range(2031, 2041)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="C6EFCE")


def _weights(n: pypsa.Network) -> pd.Series:
    return n.snapshot_weightings.generators


def _weighted_sum(df: pd.DataFrame, weights: pd.Series) -> float:
    if df is None or getattr(df, "empty", True):
        return 0.0
    return float(df.multiply(weights, axis=0).sum().sum())


def costed_dso_index(n: pypsa.Network) -> pd.Index:
    """Distribution-grid assets that carry overnight CAPEX.

    Excludes reverse-direction twins and data-centre site/demand links
    (same carrier, ``capital_cost == 0``).
    """
    dso = n.links.carrier.eq("electricity distribution grid")
    costed = n.links.capital_cost.fillna(0.0) > 0.0
    return n.links.index[dso & costed]


def de_costed_dso_index(n: pypsa.Network) -> pd.Index:
    idx = costed_dso_index(n)
    country = n.buses.reindex(n.links.loc[idx, "bus0"])["country"]
    return idx[country.fillna("").values == "DE"]


def _stock_and_new(
    df: pd.DataFrame, nom: str, opt: str, ext_col: str
) -> tuple[pd.Series, pd.Series]:
    installed = df[nom].fillna(0.0)
    optimal = df[opt].fillna(installed) if opt in df.columns else installed
    stock = optimal
    new = (optimal - installed).clip(lower=0.0)
    return stock, new


def _link_ts(n: pypsa.Network, names, side: str) -> pd.Series:
    if len(names) == 0:
        return pd.Series(0.0, index=n.snapshots)
    if side == "p1_pos":
        p = n.links_t.p1.reindex(columns=names, fill_value=0.0).clip(lower=0.0)
    else:
        p = n.links_t.p0.reindex(columns=names, fill_value=0.0).clip(lower=0.0)
    return p.sum(axis=1)


def _de_links(n: pypsa.Network, substr: str, bus_col: str) -> pd.Index:
    mask = n.links.carrier.str.contains(substr, case=False, na=False)
    buses = n.links.loc[mask, bus_col]
    country = n.buses.reindex(buses)["country"]
    return n.links.index[mask][country.fillna("").values == "DE"]


def extract_germany(n: pypsa.Network, run: str, year: int) -> dict:
    """Physical quantities for Germany only. No PyPSA euros used as CAPEX."""
    meta = RUN_META[run]
    scenario = meta["label"]
    weights = _weights(n)

    de_ac = n.buses.index[n.buses.country.eq("DE") & n.buses.carrier.eq("AC")]
    de_lv = n.buses.index[n.buses.country.eq("DE") & n.buses.carrier.eq("low voltage")]
    de_all = n.buses.index[n.buses.country.eq("DE")]

    loads = n.loads[n.loads.bus.isin(de_all)]
    p_loads = n.loads_t.p if not n.loads_t.p.empty else n.loads_t.p_set

    def load_twh(idx) -> float:
        if len(idx) == 0:
            return 0.0
        return _weighted_sum(p_loads.reindex(columns=idx, fill_value=0.0), weights) / 1e6

    def ts_load(idx) -> pd.Series:
        if len(idx) == 0:
            return pd.Series(0.0, index=n.snapshots)
        return p_loads.reindex(columns=idx, fill_value=0.0).sum(axis=1)

    conv_idx = loads.index[
        loads.bus.isin(de_lv)
        & loads.carrier.eq("electricity")
        & ~loads.index.str.contains("data center")
    ]
    dc_idx = loads.index[loads.index.str.contains("data center")]
    ind_idx = loads.index[loads.carrier.eq("industry electricity")]
    agri_idx = loads.index[
        loads.carrier.eq("agriculture electricity")
        | loads.carrier.eq("agriculture machinery electric")
    ]

    conv_twh = load_twh(conv_idx)
    dc_twh = load_twh(dc_idx)
    ind_twh = load_twh(ind_idx)
    agri_twh = load_twh(agri_idx)

    hp_names = _de_links(n, "heat pump", "bus1")
    rh_names = _de_links(n, "resistive heater", "bus0")
    bev_names = _de_links(n, "BEV charger", "bus0")
    ely_names = _de_links(n, "H2 Electrolysis", "bus0")

    hp_ts = _link_ts(n, hp_names, "p1_pos")
    rh_ts = _link_ts(n, rh_names, "p0")
    bev_ts = _link_ts(n, bev_names, "p0")
    ely_ts = _link_ts(n, ely_names, "p0")

    hp_twh = float(hp_ts.multiply(weights).sum()) / 1e6
    rh_twh = float(rh_ts.multiply(weights).sum()) / 1e6
    bev_twh = float(bev_ts.multiply(weights).sum()) / 1e6
    ely_twh = float(ely_ts.multiply(weights).sum()) / 1e6

    gens = n.generators
    shed_mask = gens.carrier.str.contains("load", case=False, na=False) | gens.index.str.contains(
        "load shedding", case=False
    )
    shed_lv = gens.index[shed_mask & gens.bus.isin(de_lv)]
    shed_ev = gens.index[
        shed_mask
        & gens.bus.map(n.buses.carrier).fillna("").str.contains("EV battery")
        & gens.bus.map(n.buses.country).eq("DE")
    ]

    def gen_twh(idx) -> float:
        if len(idx) == 0 or n.generators_t.p.empty:
            return 0.0
        p = n.generators_t.p.reindex(columns=idx, fill_value=0.0).clip(lower=0.0)
        return _weighted_sum(p, weights) / 1e6

    shed_lv_twh = gen_twh(shed_lv)
    shed_ev_twh = gen_twh(shed_ev)

    rt = n.generators.index[
        n.generators.carrier.eq("solar rooftop") & n.generators.bus.isin(de_lv)
    ]
    rt_twh = gen_twh(rt)
    rt_gw = (
        float(n.generators.loc[rt, "p_nom_opt"].fillna(n.generators.loc[rt, "p_nom"]).sum()) / 1e3
        if len(rt)
        else 0.0
    )
    rt_max = (
        float(n.generators.loc[rt, "p_nom_max"].replace([np.inf], np.nan).sum()) / 1e3
        if len(rt)
        else 0.0
    )

    enduse_twh = conv_twh + dc_twh + ind_twh + agri_twh + hp_twh + rh_twh + bev_twh
    served_twh = enduse_twh - shed_lv_twh
    gross_twh = enduse_twh + ely_twh

    enduse_ts = (
        ts_load(conv_idx)
        + ts_load(dc_idx)
        + ts_load(ind_idx)
        + ts_load(agri_idx)
        + hp_ts
        + rh_ts
        + bev_ts
    )
    peak_enduse_gw = float(enduse_ts.max()) / 1e3
    peak_enduse_when = str(enduse_ts.idxmax())

    dso_de = de_costed_dso_index(n)
    dg = n.links.loc[dso_de]
    stock, new = _stock_and_new(dg, "p_nom", "p_nom_opt", "p_nom_extendable")
    dso_stock_gw = float(stock.sum()) / 1e3
    dso_new_gw = float(new.sum()) / 1e3
    dso_new_ann_bn = float((new * dg["capital_cost"]).sum()) / 1e9
    dso_stock_ann_bn = float((stock * dg["capital_cost"]).sum()) / 1e9

    dso_all = n.links.loc[costed_dso_index(n)]
    stock_e, new_e = _stock_and_new(dso_all, "p_nom", "p_nom_opt", "p_nom_extendable")
    eu_new_ann_bn = float((new_e * dso_all["capital_cost"]).sum()) / 1e9
    eu_stock_gw = float(stock_e.sum()) / 1e3

    p1 = n.links_t.p1.reindex(columns=dso_de, fill_value=0.0)
    delivered = (-p1).clip(lower=0.0)
    dso_twh = _weighted_sum(delivered, weights) / 1e6
    dso_flow_ts = delivered.sum(axis=1)
    peak_dso_gw = float(dso_flow_ts.max()) / 1e3
    peak_dso_when = str(dso_flow_ts.idxmax())
    util = 100.0 * peak_dso_gw / dso_stock_gw if dso_stock_gw else float("nan")
    headroom_gw = dso_stock_gw - peak_dso_gw

    budget = EU_DSO_BUDGET_BN.get((scenario, int(year)))
    if budget is None:
        slack_bn = float("nan")
        binding = float("nan")
    else:
        slack_bn = budget - eu_new_ann_bn
        binding = abs(slack_bn) < 0.05

    lines = n.lines.copy()
    c0 = n.buses.reindex(lines["bus0"])["country"].values
    c1 = n.buses.reindex(lines["bus1"])["country"].values
    internal = (c0 == "DE") & (c1 == "DE")
    xborder = (c0 == "DE") ^ (c1 == "DE")
    stock_l, new_l = _stock_and_new(lines, "s_nom", "s_nom_opt", "s_nom_extendable")
    length = lines["length"].fillna(0.0)

    def gwkm(cap: pd.Series, mask, share: float) -> float:
        return float((cap[mask] * length[mask] * share).sum()) / 1e3

    ac_vol_stock = gwkm(stock_l, internal, 1.0) + gwkm(stock_l, xborder, 0.5)
    ac_vol_new = gwkm(new_l, internal, 1.0) + gwkm(new_l, xborder, 0.5)
    ac_ann_bn = (
        float((stock_l[internal] * lines.loc[internal, "capital_cost"]).sum())
        + 0.5 * float((stock_l[xborder] * lines.loc[xborder, "capital_cost"]).sum())
    ) / 1e9
    ac_new_ann_bn = (
        float((new_l[internal] * lines.loc[internal, "capital_cost"]).sum())
        + 0.5 * float((new_l[xborder] * lines.loc[xborder, "capital_cost"]).sum())
    ) / 1e9

    dc = n.links[n.links.carrier.eq("DC")].copy()
    if dc.empty:
        dc_vol_stock = dc_vol_new = dc_ann_bn = dc_new_ann_bn = 0.0
        dc_int_stock = dc_int_new = dc_xb_stock = dc_xb_new = 0.0
    else:
        d0 = n.buses.reindex(dc["bus0"])["country"].values
        d1 = n.buses.reindex(dc["bus1"])["country"].values
        d_int = (d0 == "DE") & (d1 == "DE")
        d_xb = (d0 == "DE") ^ (d1 == "DE")
        stock_d, new_d = _stock_and_new(dc, "p_nom", "p_nom_opt", "p_nom_extendable")
        dlen = dc["length"].fillna(0.0) if "length" in dc.columns else pd.Series(0.0, index=dc.index)
        dc_int_stock = float(stock_d[d_int].sum()) / 1e3
        dc_int_new = float(new_d[d_int].sum()) / 1e3
        dc_xb_stock = float(stock_d[d_xb].sum()) / 1e3
        dc_xb_new = float(new_d[d_xb].sum()) / 1e3
        dc_vol_stock = float(
            (stock_d[d_int] * dlen[d_int]).sum() + 0.5 * (stock_d[d_xb] * dlen[d_xb]).sum()
        ) / 1e3
        dc_vol_new = float(
            (new_d[d_int] * dlen[d_int]).sum() + 0.5 * (new_d[d_xb] * dlen[d_xb]).sum()
        ) / 1e3
        dc_ann_bn = (
            float((stock_d[d_int] * dc.loc[d_int, "capital_cost"]).sum())
            + 0.5 * float((stock_d[d_xb] * dc.loc[d_xb, "capital_cost"]).sum())
        ) / 1e9
        dc_new_ann_bn = (
            float((new_d[d_int] * dc.loc[d_int, "capital_cost"]).sum())
            + 0.5 * float((new_d[d_xb] * dc.loc[d_xb, "capital_cost"]).sum())
        ) / 1e9

    lv_mu = (
        float(n.global_constraints.at["lv_limit", "mu"])
        if "lv_limit" in n.global_constraints.index
        else float("nan")
    )

    return {
        "run": run,
        "scenario": scenario,
        "cap_fraction_label": meta["cap_fraction"],
        "year": int(year),
        "de_ac_nodes": int(len(de_ac)),
        "conv_lv_twh": conv_twh,
        "industry_twh": ind_twh,
        "agriculture_twh": agri_twh,
        "datacentre_twh": dc_twh,
        "heat_pump_twh": hp_twh,
        "resistive_heater_twh": rh_twh,
        "bev_twh": bev_twh,
        "electrolysis_twh": ely_twh,
        "enduse_twh": enduse_twh,
        "served_enduse_twh": served_twh,
        "gross_incl_ptx_twh": gross_twh,
        "rooftop_twh": rt_twh,
        "rooftop_gw": rt_gw,
        "rooftop_pmax_gw": rt_max,
        "lv_shed_twh": shed_lv_twh,
        "ev_shed_twh": shed_ev_twh,
        "peak_enduse_gw": peak_enduse_gw,
        "peak_enduse_when": peak_enduse_when,
        "dso_stock_gw": dso_stock_gw,
        "dso_new_gw": dso_new_gw,
        "dso_hv_lv_twh": dso_twh,
        "peak_dso_gw": peak_dso_gw,
        "peak_dso_when": peak_dso_when,
        "dso_peak_util_pct": util,
        "dso_headroom_gw": headroom_gw,
        "dso_new_ann_bn_context_only": dso_new_ann_bn,
        "dso_stock_ann_bn_context_only": dso_stock_ann_bn,
        "eu_dso_new_ann_bn": eu_new_ann_bn,
        "eu_dso_stock_gw": eu_stock_gw,
        "eu_budget_bn": budget if budget is not None else float("nan"),
        "eu_budget_slack_bn": slack_bn,
        "eu_budget_binding": binding,
        "de_share_of_eu_dso_stock": dso_stock_gw / eu_stock_gw if eu_stock_gw else float("nan"),
        "ac_internal_stock_gw": float(stock_l[internal].sum()) / 1e3,
        "ac_internal_new_gw": float(new_l[internal].sum()) / 1e3,
        "ac_xborder_stock_gw": float(stock_l[xborder].sum()) / 1e3,
        "ac_xborder_new_gw": float(new_l[xborder].sum()) / 1e3,
        "ac_vol_stock_gwkm": ac_vol_stock,
        "ac_vol_new_gwkm": ac_vol_new,
        "dc_internal_stock_gw": dc_int_stock,
        "dc_internal_new_gw": dc_int_new,
        "dc_xborder_stock_gw": dc_xb_stock,
        "dc_xborder_new_gw": dc_xb_new,
        "dc_vol_stock_gwkm": dc_vol_stock,
        "dc_vol_new_gwkm": dc_vol_new,
        "tso_vol_stock_gwkm": ac_vol_stock + dc_vol_stock,
        "tso_vol_new_gwkm": ac_vol_new + dc_vol_new,
        "tso_ann_bn_context_only": ac_ann_bn + dc_ann_bn,
        "tso_new_ann_bn_context_only": ac_new_ann_bn + dc_new_ann_bn,
        "lv_limit_mu": lv_mu,
        "dso_budget_dual": "not stored in solved .nc",
    }


def discover_networks(results_root: Path) -> list[tuple[str, int, Path]]:
    found = []
    for run, meta in RUN_META.items():
        networks = results_root / run / "networks"
        if not networks.is_dir():
            continue
        for year in (2030, 2040):
            path = networks / f"base_s_37___{year}.nc"
            if path.exists():
                found.append((run, year, path))
            else:
                logger.warning("Missing %s", path)
    return found


def collect_extracts(results_root: Path) -> pd.DataFrame:
    rows = []
    for run, year, path in discover_networks(results_root):
        logger.info("Reading %s", path)
        n = pypsa.Network(str(path))
        rows.append(extract_germany(n, run, year))
        del n
    df = pd.DataFrame(rows)
    return df.sort_values(["year", "cap_fraction_label"], ascending=[True, False]).reset_index(
        drop=True
    )


def excel_params_from_original(original: Path) -> dict:
    """Read financial baseline from the German workbook (inspected, not assumed)."""
    wb = openpyxl.load_workbook(original, data_only=False)
    ws = wb["Parameters"]
    rab = wb["RAB & Cost Build-up"]

    inv_b = {}
    demand = {"A": {}, "B": {}, "C": {}}
    opex = {}
    orig_tariff = {"A": {}, "B": {}, "C": {}}
    # Row 4 = 2023. Scenario B inv=J, demand=P, opex=N, sys avg=Q.
    for row in range(4, rab.max_row + 1):
        year = rab.cell(row, 1).value
        if not isinstance(year, (int, float)):
            continue
        year = int(year)
        inv_b[year] = float(rab.cell(row, 10).value)  # J
        opex[year] = float(rab.cell(row, 14).value)  # N
        demand["A"][year] = float(rab.cell(row, 8).value)  # H
        demand["B"][year] = float(rab.cell(row, 16).value)  # P
        demand["C"][year] = float(rab.cell(row, 24).value)  # X
        orig_tariff["A"][year] = float(rab.cell(row, 9).value)  # I ct/kWh
        orig_tariff["B"][year] = float(rab.cell(row, 17).value)  # Q
        orig_tariff["C"][year] = float(rab.cell(row, 25).value)  # Y

    return {
        "wacc_early": float(ws["B5"].value),
        "wacc_late": float(ws["B6"].value),
        "wacc_switch_year": 2029,
        "start_rab": float(ws["B9"].value),
        "dep_rate": float(ws["B10"].value),
        "start_year": 2023,
        "inv_base": inv_b,
        "opex": opex,
        "demand_excel": demand,
        "orig_tariff_ct": orig_tariff,
        "band_ig": float(ws["B37"].value),
        "band_if": float(ws["B38"].value),
        "band_ie": float(ws["B39"].value),
        "band_id": float(ws["B40"].value),
        "band_ic": float(ws["B41"].value),
        "tso_share": TSO_SHARE,
        "dso_share": DSO_SHARE,
    }


def wacc_for_year(params: dict, year: int) -> float:
    return params["wacc_early"] if year < params["wacc_switch_year"] else params["wacc_late"]


def roll_rab(
    params: dict,
    inv_fn,
    demand_fn,
    years=range(2023, 2041),
) -> pd.DataFrame:
    """Replicate the German Excel identities: WACC on opening RAB, Dep = rate × opening."""
    opening = params["start_rab"]
    rows = []
    for year in years:
        inv = float(inv_fn(year))
        dep = params["dep_rate"] * opening
        wacc = wacc_for_year(params, year)
        ox = float(params["opex"][year])
        cost = wacc * opening + dep + ox
        dem = float(demand_fn(year))
        rows.append(
            {
                "year": year,
                "inv_bn": inv,
                "depreciation_bn": dep,
                "opening_rab_bn": opening,
                "closing_rab_bn": opening + inv - dep,
                "wacc_times_rab_bn": wacc * opening,
                "opex_bn": ox,
                "total_cost_bn": cost,
                "demand_twh": dem,
                "system_avg_ct_per_kwh": 100.0 * cost / dem if dem else float("nan"),
                "system_avg_eur_per_mwh": 1000.0 * cost / dem if dem else float("nan"),
                "system_avg_eur_per_kwh": cost / dem if dem else float("nan"),
            }
        )
        opening = opening + inv - dep
    return pd.DataFrame(rows)


def _get(extracts: pd.DataFrame, scenario: str, year: int, col: str) -> float:
    m = extracts[(extracts["scenario"] == scenario) & (extracts["year"] == year)]
    if m.empty:
        raise ValueError(f"Missing PyPSA extract for {scenario} {year}")
    return float(m.iloc[0][col])


def ratio_vs_build(extracts: pd.DataFrame, scenario: str, year: int, col: str) -> float:
    return _get(extracts, scenario, year, col) / _get(extracts, "Build", year, col)


def demand_fn_pypsa(extracts: pd.DataFrame, params: dict, scenario: str):
    d2030 = _get(extracts, scenario, 2030, "enduse_twh")
    d2040 = _get(extracts, scenario, 2040, "enduse_twh")
    d2025 = params["demand_excel"]["B"][2025]

    def dem(year: int) -> float:
        if year <= HISTORICAL_THROUGH:
            return params["demand_excel"]["B"][year]
        if year == 2030:
            return d2030
        if year == 2040:
            return d2040
        if HISTORICAL_THROUGH < year < 2030:
            return d2025 + (d2030 - d2025) * (year - HISTORICAL_THROUGH) / (
                2030 - HISTORICAL_THROUGH
            )
        if 2030 < year < 2040:
            return d2030 + (d2040 - d2030) * (year - 2030) / 10.0
        return d2040

    return dem


def inv_fn_scaled(params: dict, f2030: float, f2040: float):
    def inv(year: int) -> float:
        base = params["inv_base"][year]
        if year <= HISTORICAL_THROUGH:
            return base
        if year <= 2030:
            return base * f2030
        return base * f2040

    return inv


def inv_fn_split(
    params: dict,
    dso_f2030: float,
    dso_f2040: float,
    tso_f2030: float,
    tso_f2040: float,
):
    def inv(year: int) -> float:
        base = params["inv_base"][year]
        if year <= HISTORICAL_THROUGH:
            return base
        if year <= 2030:
            return base * (DSO_SHARE * dso_f2030 + TSO_SHARE * tso_f2030)
        return base * (DSO_SHARE * dso_f2040 + TSO_SHARE * tso_f2040)

    return inv


CASES = [
    (
        "A_demand_scaled",
        "Case A — Excel method: Inv × (end-use_s / end-use_Build)",
    ),
    (
        "B_denominator_only",
        "Case B — Excel B investment unchanged; PyPSA end-use denominator only",
    ),
    (
        "C_dso_capacity_all_inv",
        "Case C — all Inv_base × (German DSO stock_s / stock_Build)",
    ),
    (
        "D_peak_flow_all_inv",
        "Case D — all Inv_base × (German peak HV→LV_s / peak_Build)",
    ),
    (
        "E_tso_dso_split",
        "Case E — VNB share × DSO peak; ÜNB share × TSO stock GW-km (recommended)",
    ),
    (
        "E_dso_only_tso_nep",
        "Sensitivity — VNB share × DSO peak; ÜNB share left at Excel B (NEP TSO continues)",
    ),
    (
        "E_tso_new_volume",
        "Sensitivity — VNB × DSO peak; ÜNB × TSO *new* GW-km (not recommended)",
    ),
]


def build_all_cases(extracts: pd.DataFrame, params: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Snapshot (2030/2040) comparison and annual RAB paths."""
    scenarios = [m["label"] for m in RUN_META.values()]
    annual_frames = []
    snap_rows = []

    orig_rows = []
    for scen_key, label in (("A", "Excel A (NEP low)"), ("B", "Excel B (NEP base)"), ("C", "Excel C (NEP high)")):
        for year in (2030, 2040):
            orig_rows.append(
                {
                    "scenario": label,
                    "year": year,
                    "case": "original_excel",
                    "pypsa_enduse_twh": params["demand_excel"][scen_key][year],
                    "pypsa_dso_hv_lv_twh": float("nan"),
                    "peak_dso_gw": float("nan"),
                    "dso_stock_gw": float("nan"),
                    "dso_vs_build_pct": float("nan"),
                    "dso_new_vs_build_pct": float("nan"),
                    "tso_stock_vs_build_pct": float("nan"),
                    "tso_new_vs_build_pct": float("nan"),
                    "inv_bn": params["inv_base"][year] if scen_key == "B" else float("nan"),
                    "total_cost_bn": float("nan"),
                    "system_avg_ct_per_kwh": params["orig_tariff_ct"][scen_key][year],
                    "ig_eur_per_kwh": params["orig_tariff_ct"][scen_key][year]
                    / 100.0
                    * params["band_ig"],
                    "ic_eur_per_kwh": params["orig_tariff_ct"][scen_key][year]
                    / 100.0
                    * params["band_ic"],
                    "notes": "Workbook NEP demand scenario; not a PyPSA grid-cap case",
                }
            )
    snap_rows.extend(orig_rows)

    for scenario in scenarios:
        dem = demand_fn_pypsa(extracts, params, scenario)
        f_dem_30 = ratio_vs_build(extracts, scenario, 2030, "enduse_twh")
        f_dem_40 = ratio_vs_build(extracts, scenario, 2040, "enduse_twh")
        f_cap_30 = ratio_vs_build(extracts, scenario, 2030, "dso_stock_gw")
        f_cap_40 = ratio_vs_build(extracts, scenario, 2040, "dso_stock_gw")
        f_peak_30 = ratio_vs_build(extracts, scenario, 2030, "peak_dso_gw")
        f_peak_40 = ratio_vs_build(extracts, scenario, 2040, "peak_dso_gw")
        f_tso_s_30 = ratio_vs_build(extracts, scenario, 2030, "tso_vol_stock_gwkm")
        f_tso_s_40 = ratio_vs_build(extracts, scenario, 2040, "tso_vol_stock_gwkm")
        f_tso_n_30 = ratio_vs_build(extracts, scenario, 2030, "tso_vol_new_gwkm")
        f_tso_n_40 = ratio_vs_build(extracts, scenario, 2040, "tso_vol_new_gwkm")

        case_inv = {
            "A_demand_scaled": inv_fn_scaled(params, f_dem_30, f_dem_40),
            "B_denominator_only": lambda y, _p=params: _p["inv_base"][y],
            "C_dso_capacity_all_inv": inv_fn_scaled(params, f_cap_30, f_cap_40),
            "D_peak_flow_all_inv": inv_fn_scaled(params, f_peak_30, f_peak_40),
            "E_tso_dso_split": inv_fn_split(
                params, f_peak_30, f_peak_40, f_tso_s_30, f_tso_s_40
            ),
            "E_dso_only_tso_nep": inv_fn_split(params, f_peak_30, f_peak_40, 1.0, 1.0),
            "E_tso_new_volume": inv_fn_split(
                params, f_peak_30, f_peak_40, f_tso_n_30, f_tso_n_40
            ),
        }

        for case_id, case_label in CASES:
            sim = roll_rab(params, case_inv[case_id], dem)
            sim.insert(0, "scenario", scenario)
            sim.insert(1, "case", case_id)
            sim.insert(2, "case_label", case_label)
            annual_frames.append(sim)

            for year in (2030, 2040):
                srow = sim[sim["year"] == year].iloc[0]
                snap_rows.append(
                    {
                        "scenario": scenario,
                        "year": year,
                        "case": case_id,
                        "case_label": case_label,
                        "pypsa_enduse_twh": _get(extracts, scenario, year, "enduse_twh"),
                        "pypsa_dso_hv_lv_twh": _get(
                            extracts, scenario, year, "dso_hv_lv_twh"
                        ),
                        "peak_dso_gw": _get(extracts, scenario, year, "peak_dso_gw"),
                        "dso_stock_gw": _get(extracts, scenario, year, "dso_stock_gw"),
                        "dso_vs_build_pct": 100.0
                        * ratio_vs_build(extracts, scenario, year, "dso_stock_gw"),
                        "dso_new_vs_build_pct": 100.0
                        * ratio_vs_build(extracts, scenario, year, "dso_new_gw"),
                        "peak_dso_vs_build_pct": 100.0
                        * ratio_vs_build(extracts, scenario, year, "peak_dso_gw"),
                        "tso_stock_vs_build_pct": 100.0
                        * ratio_vs_build(
                            extracts, scenario, year, "tso_vol_stock_gwkm"
                        ),
                        "tso_new_vs_build_pct": 100.0
                        * ratio_vs_build(extracts, scenario, year, "tso_vol_new_gwkm"),
                        "enduse_vs_build_pct": 100.0
                        * ratio_vs_build(extracts, scenario, year, "enduse_twh"),
                        "inv_bn": srow["inv_bn"],
                        "closing_rab_bn": srow["closing_rab_bn"],
                        "total_cost_bn": srow["total_cost_bn"],
                        "system_avg_ct_per_kwh": srow["system_avg_ct_per_kwh"],
                        "system_avg_eur_per_mwh": srow["system_avg_eur_per_mwh"],
                        "ig_eur_per_kwh": srow["system_avg_eur_per_kwh"]
                        * params["band_ig"],
                        "ic_eur_per_kwh": srow["system_avg_eur_per_kwh"]
                        * params["band_ic"],
                        "dso_peak_util_pct": _get(
                            extracts, scenario, year, "dso_peak_util_pct"
                        ),
                        "lv_shed_twh": _get(extracts, scenario, year, "lv_shed_twh"),
                        "notes": "",
                    }
                )

    annual = pd.concat(annual_frames, ignore_index=True)
    snap = pd.DataFrame(snap_rows)
    return snap, annual


def _style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, freeze: bool = False) -> None:
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for j, val in enumerate(row, 1):
            if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                val = None
            ws.cell(i, j, val)
        if i == start_row:
            _style_header(ws, i, len(row))
    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"


def copy_and_fill_workbook(
    original: Path,
    dest: Path,
    extracts: pd.DataFrame,
    params: dict,
    snap: pd.DataFrame,
    annual: pd.DataFrame,
) -> None:
    shutil.copy2(original, dest)
    wb = openpyxl.load_workbook(dest)

    # Keep every original sheet. Prepend PyPSA sheets.
    _write_cover_sheet(wb, params)
    _write_inputs_sheet(wb, extracts)
    _write_mapping_sheet(wb)
    _write_comparison_sheet(wb, snap)
    _write_recommended_rab_sheet(wb, annual)
    _write_methodology_sheet(wb)

    wb.save(dest)
    logger.info("Wrote %s", dest)


def _write_cover_sheet(wb: openpyxl.Workbook, params: dict) -> None:
    if "PyPSA_Cover" in wb.sheetnames:
        del wb["PyPSA_Cover"]
    ws = wb.create_sheet("PyPSA_Cover", 0)
    lines = [
        ("German network-tariff model — PyPSA physical overlay",),
        ("Original workbook is unchanged. This file is a copy.",),
        ("",),
        ("Financial baseline (Excel, not PyPSA)",),
        ("Opening RAB 2023 (bn €)", params["start_rab"]),
        ("WACC to 2028 / from 2029", f"{params['wacc_early']:.4f} / {params['wacc_late']:.4f}"),
        ("Depreciation rate", params["dep_rate"]),
        ("OPEX path", "Excel Scenario B (15→27 bn, same in every PyPSA case)"),
        ("Inv_base", "Excel Scenario B IMK schedule"),
        ("VNB / ÜNB split of Inv_base", f"{DSO_SHARE:.4f} / {TSO_SHARE:.4f} (14.4 / 19.8 of 34.2)"),
        ("",),
        ("PyPSA used for",),
        ("German DSO stock and peak HV→LV flow", "investment scaling (VNB share)"),
        ("German TSO stock GW-km (internal 100%, cross-border 50%)", "investment scaling (ÜNB share)"),
        ("German end-use electricity TWh", "tariff denominator"),
        ("",),
        ("PyPSA NOT used for",),
        ("Absolute DSO/TSO CAPEX euros", "too aggregated"),
        ("Existing RAB, WACC, OPEX, historical investment", "Excel / BNetzA / BDEW / IMK"),
        ("85/75/65/50% labels as German capacity", "those are EU-wide DSO budget caps"),
        ("",),
        ("Recommended case", "E_tso_dso_split (sheet PyPSA_RAB_recommended)"),
        ("Comparison of all cases", "sheet PyPSA_Comparison"),
    ]
    for i, row in enumerate(lines, 1):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
        if i == 1:
            ws.cell(i, 1).font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 70


def _write_inputs_sheet(wb: openpyxl.Workbook, extracts: pd.DataFrame) -> None:
    if "Inputs_PyPSA_DE" in wb.sheetnames:
        del wb["Inputs_PyPSA_DE"]
    ws = wb.create_sheet("Inputs_PyPSA_DE", 1)
    ws["A1"] = (
        "German physical quantities from solved PyPSA networks. "
        "DSO assets: costed links (carrier 'electricity distribution grid' and capital_cost>0) "
        "at DE AC nodes. Absolute overnight CAPEX is context-only and is not written into Inv."
    )
    _write_df(ws, extracts, start_row=3, freeze=True)


def _write_mapping_sheet(wb: openpyxl.Workbook) -> None:
    if "PyPSA_Mapping" in wb.sheetnames:
        del wb["PyPSA_Mapping"]
    ws = wb.create_sheet("PyPSA_Mapping", 2)
    rows = [
        ("Sheet", "What changed", "Source", "Notes"),
        (
            "Parameters / RAB & Cost Build-up / Tariff Results / Scenario Comparison / Methodology Note",
            "Untouched copies of the original model",
            "german_network_tariff_model_v3.xlsx",
            "NEP A/B/C demand scenarios remain the Excel-only result.",
        ),
        (
            "PyPSA_Comparison",
            "Cases A–E system-average tariffs for Build/85/75/65/50% in 2030 and 2040",
            "Excel B financial identities + German PyPSA ratios",
            "Case E is the publication recommendation.",
        ),
        (
            "PyPSA_RAB_recommended",
            "Year-by-year RAB for Case E (VNB×peak DSO, ÜNB×TSO stock)",
            "Same identities as original RAB sheet",
            "2023–2025 investment is Excel B historical. 2026–30 uses 2030 ratios; 2031–40 uses 2040 ratios.",
        ),
        (
            "Inputs_PyPSA_DE",
            "Raw German extract",
            "solved .nc, 3 DE nodes",
            "Industry/households/HP/BEV/data centres are on LV. No HöS-connected industry in the model.",
        ),
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
        if i == 1:
            _style_header(ws, 1, 4)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 70


def _write_comparison_sheet(wb: openpyxl.Workbook, snap: pd.DataFrame) -> None:
    if "PyPSA_Comparison" in wb.sheetnames:
        del wb["PyPSA_Comparison"]
    ws = wb.create_sheet("PyPSA_Comparison", 3)

    # Wide table: one row per scenario-year with case tariffs as columns.
    core = snap[snap["case"] != "original_excel"].copy()
    wide_rows = []
    for (scenario, year), g in core.groupby(["scenario", "year"], sort=False):
        row = {
            "scenario": scenario,
            "year": year,
            "pypsa_enduse_twh": g["pypsa_enduse_twh"].iloc[0],
            "peak_dso_gw": g["peak_dso_gw"].iloc[0],
            "dso_stock_gw": g["dso_stock_gw"].iloc[0],
            "dso_vs_build_pct": g["dso_vs_build_pct"].iloc[0],
            "tso_stock_vs_build_pct": g["tso_stock_vs_build_pct"].iloc[0],
            "tso_new_vs_build_pct": g["tso_new_vs_build_pct"].iloc[0],
            "enduse_vs_build_pct": g["enduse_vs_build_pct"].iloc[0],
        }
        for case_id, _label in CASES:
            sub = g[g["case"] == case_id]
            row[case_id] = float(sub["system_avg_ct_per_kwh"].iloc[0]) if not sub.empty else None
        wide_rows.append(row)
    wide = pd.DataFrame(wide_rows)

    orig = snap[snap["case"] == "original_excel"][
        ["scenario", "year", "pypsa_enduse_twh", "system_avg_ct_per_kwh"]
    ].rename(
        columns={
            "pypsa_enduse_twh": "nep_demand_twh",
            "system_avg_ct_per_kwh": "original_excel_ct_per_kwh",
        }
    )

    ws["A1"] = (
        "System-average network tariff (ct/kWh). "
        "Excel A/B/C are NEP demand scenarios (original methodology). "
        "Build/85/75/65/50% use Excel B finance + PyPSA German physics. "
        "Case E is recommended. Cases C and D are empirically identical (DSO utilisation 97% in every run)."
    )
    ws.merge_cells("A1:P1")
    ws["A3"] = "Original Excel NEP scenarios (ct/kWh, from the copied workbook)"
    _write_df(ws, orig, start_row=4)

    start = 4 + len(orig) + 3
    ws.cell(start - 1, 1, "PyPSA grid-cap scenarios (ct/kWh system average)")
    _write_df(ws, wide, start_row=start)

    # Full long table below
    long_start = start + len(wide) + 4
    ws.cell(long_start - 1, 1, "Long table (all cases, including band IG/IC)")
    _write_df(ws, snap, start_row=long_start)


def _write_recommended_rab_sheet(wb: openpyxl.Workbook, annual: pd.DataFrame) -> None:
    if "PyPSA_RAB_recommended" in wb.sheetnames:
        del wb["PyPSA_RAB_recommended"]
    ws = wb.create_sheet("PyPSA_RAB_recommended", 4)
    rec = annual[annual["case"] == "E_tso_dso_split"].copy()
    ws["A1"] = (
        "Case E (recommended): Inv_t = Inv_B,t × [0.4211 × (peak DSO_s / peak DSO_Build) "
        "+ 0.5789 × (TSO stock GW-km_s / TSO stock_Build)]. "
        "Demand = German PyPSA end-use TWh. OPEX/WACC/depreciation from Excel B. "
        "Yellow years 2023–2025 are historical (unscaled)."
    )
    ws.merge_cells("A1:H1")

    scenarios = [m["label"] for m in RUN_META.values()]
    years = list(range(2023, 2041))
    blocks = [
        ("Inv (bn €)", "inv_bn"),
        ("Closing RAB (bn €)", "closing_rab_bn"),
        ("Total cost (bn €)", "total_cost_bn"),
        ("Demand (TWh)", "demand_twh"),
        ("Sys avg (ct/kWh)", "system_avg_ct_per_kwh"),
    ]
    col = 2
    colmap = {}  # (block, scenario) -> col
    ws.cell(3, 1, "Year")
    c = 2
    for title, _field in blocks:
        ws.cell(2, c, title)
        for sc in scenarios:
            ws.cell(3, c, sc)
            colmap[(title, sc)] = c
            c += 1

    _style_header(ws, 3, c - 1)
    for i, year in enumerate(years, 4):
        ws.cell(i, 1, year)
        if year <= HISTORICAL_THROUGH:
            ws.cell(i, 1).fill = YELLOW
        if year in (2030, 2040):
            ws.cell(i, 1).fill = GREEN
        for title, field in blocks:
            for sc in scenarios:
                sub = rec[(rec["scenario"] == sc) & (rec["year"] == year)]
                val = float(sub[field].iloc[0]) if not sub.empty else None
                ws.cell(i, colmap[(title, sc)], val)

    ws.cell(24, 1, "Green = PyPSA snapshot years (2030, 2040). Yellow = historical benchmark.")
    ws.freeze_panes = "B4"


def _write_methodology_sheet(wb: openpyxl.Workbook) -> None:
    if "PyPSA_Methodology" in wb.sheetnames:
        del wb["PyPSA_Methodology"]
    ws = wb.create_sheet("PyPSA_Methodology", 5)
    for i, line in enumerate(METHODOLOGY.splitlines(), 1):
        ws.cell(i, 1, line)
    ws.column_dimensions["A"].width = 120


METHODOLOGY = """# German network tariffs: Excel finance + PyPSA physics

## Rule

The original workbook (`german_network_tariff_model_v3.xlsx`) is not modified.
This copy keeps every original sheet. PyPSA absolute grid CAPEX is never used
as German investment.

Excel supplies the realistic financial/regulatory baseline:
opening RAB 186 bn € (2023), WACC 5.04% then 5.50%, depreciation 3%/yr of
opening RAB, OPEX 15→27 bn € (scenario-independent), Inv_base from IMK
(Scenario B).

PyPSA supplies endogenous *German* physical ratios only.

## Excel identities (inspected)

    Tariff_t = (WACC_t × RAB_t + Dep_t + OPEX_t) / Demand_t
    RAB_{t+1} = RAB_t + Inv_t − Dep_t
    Dep_t = 0.03 × opening RAB_t
    Inv_s(t) = Inv_B(t) × Demand_s(t) / Demand_B(t)   # original NEP A/C method

Original Excel A/B/C are NEP 2037/2045 Version 2025 *demand* scenarios
(967 / 1,179 / 1,351 TWh Brutto in 2045). They are not grid-cap scenarios.

Documented TSO/DSO split of the 34 bn peak year: ÜNB 19.8 + VNB 14.4.
Shares used here are 19.8/34.2 and 14.4/34.2.

Band allocation (IG 0.22 … IC 1.30) is a multiplier on the *same* system
average. It is not a voltage-specific cost pool. PyPSA cannot assign
customers to HöS vs NS: industry electricity sits on LV.

## PyPSA geography and assets

Germany = 3 AC nodes (DE2 0/1/2) in the 37-cluster networks.
Costed DSO = carrier 'electricity distribution grid' AND capital_cost > 0.
Reverse twins and data-centre site links are excluded.

The 85/75/65/50% names are a *Europe-wide* cap on new DSO annualised
investment. They are not German capacity findings. German realised DSO
stock vs Build is endogenous (2030: 94.4 / 91.0 / 79.8 / 59.8%;
2040: 83.7 / 75.4 / 65.3 / 52.7%). Peak HV→LV ratios match stock ratios
because utilisation is 97% in every run (DSO is sized to peak).

The DSO-budget dual is not stored in the solved .nc files. The EU cap is
binding (slack ≈ 0) in every inaction run.

Cross-border AC/DC volume is allocated 50% to Germany.

## Denominator

Excel recovers TSO+DSO from total German kWh. The matching PyPSA quantity
is German *end-use electricity*:
conventional LV + industry + agriculture + data centres + heat pumps
+ resistive heaters + BEV chargers.

Not used as the system-average denominator:
- HV→LV DSO flow (excludes rooftop self-consumption; Excel is TSO+DSO)
- Gross including electrolysis (50% 2040 PtX jumps to ~270 TWh and would
  inflate the billing base; NEP Brutto includes PtX, StromNEV bands do not)

## Cases

A  Excel method on PyPSA scenarios: Inv × (end-use_s / end-use_Build).
B  Excel B investment unchanged; only the denominator is PyPSA end-use.
C  All Inv_base × German DSO stock ratio. Over-scales TSO. Not recommended.
D  All Inv_base × German peak HV→LV ratio. Empirically identical to C.
E  Recommended. VNB share × peak DSO; ÜNB share × TSO *stock* GW-km.
   Historical 2023–2025 unscaled.

Sensitivities on the E sheet set: TSO left at NEP (E_dso_only_tso_nep);
TSO scaled by *new* GW-km (E_tso_new_volume; 50% 2030 new TSO is ~7% of
Build and is not a realistic replacement programme).

## What is exogenous vs endogenous

Exogenous: EU DSO budget labels, Excel RAB/WACC/OPEX/Inv_base, OPEX
invariance, band weights, conventional/industry/agri/data-centre TWh.

Endogenous: German realised DSO capacity and peak flow; German TSO
expansion (it *falls* with the DSO cap, it does not substitute);
HP/RH/BEV; rooftop (hits the German cap from 65% 2030); shedding
(50% 2030 only, ~20 TWh LV); electrolysis; Germany's rising share of
the scarce EU DSO pie (16% → ~19–21%).

## Publication recommendation

Use Case E. Peak is the right physical driver for DSO investment; in
these solves it equals capacity. Do not present 85/75/65/50% as German
DSO findings. Do not put PyPSA euros into the German RAB.
"""


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    p.add_argument("--results-root", type=Path, default=Path("results/2026-grid-study"))
    p.add_argument(
        "--original-xlsx",
        type=Path,
        default=Path(
            "results/2026-grid-study/distribution_tariffs/german_network_tariff_model_v3.xlsx"
        ),
    )
    p.add_argument(
        "--output-dir",
        type=Path,
        default=Path(
            "results/2026-grid-study/distribution_tariffs/german_network_tariffs"
        ),
    )
    p.add_argument(
        "--extract-csv",
        type=Path,
        default=None,
        help="Reuse a previously written pypsa_germany_extract.csv.",
    )
    return p.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    if args.extract_csv is not None:
        extracts = pd.read_csv(args.extract_csv)
        logger.info("Reused PyPSA extract from %s", args.extract_csv)
    else:
        extracts = collect_extracts(args.results_root)

    extracts_path = args.output_dir / "pypsa_germany_extract.csv"
    extracts.to_csv(extracts_path, index=False, float_format="%.6f")
    logger.info("Wrote %s", extracts_path)

    params = excel_params_from_original(args.original_xlsx)
    snap, annual = build_all_cases(extracts, params)

    snap_path = args.output_dir / "german_tariff_comparison.csv"
    snap.to_csv(snap_path, index=False, float_format="%.4f")
    logger.info("Wrote %s", snap_path)

    annual_path = args.output_dir / "german_rab_annual.csv"
    annual.to_csv(annual_path, index=False, float_format="%.4f")
    logger.info("Wrote %s", annual_path)

    md_path = args.output_dir / "methodology.md"
    md_path.write_text(METHODOLOGY)
    logger.info("Wrote %s", md_path)

    xlsx_path = args.output_dir / "german_network_tariff_model_v3_PyPSA.xlsx"
    copy_and_fill_workbook(args.original_xlsx, xlsx_path, extracts, params, snap, annual)

    # Headline table: original Excel + Case A/B/C/D/E
    show = snap[
        snap["case"].isin(
            [
                "original_excel",
                "A_demand_scaled",
                "B_denominator_only",
                "C_dso_capacity_all_inv",
                "D_peak_flow_all_inv",
                "E_tso_dso_split",
            ]
        )
    ].copy()
    wide_src = show[show["case"] != "original_excel"]
    print("\nOriginal Excel (NEP A/B/C, ct/kWh):")
    orig = show[show["case"] == "original_excel"][
        ["scenario", "year", "system_avg_ct_per_kwh"]
    ]
    print(orig.to_string(index=False, float_format=lambda x: f"{x:.2f}"))

    print("\nPyPSA-informed system average (ct/kWh):")
    cols = [
        "scenario",
        "year",
        "pypsa_enduse_twh",
        "peak_dso_gw",
        "dso_stock_gw",
        "dso_vs_build_pct",
        "tso_stock_vs_build_pct",
        "A_demand_scaled",
        "B_denominator_only",
        "C_dso_capacity_all_inv",
        "D_peak_flow_all_inv",
        "E_tso_dso_split",
    ]
    rows = []
    for (scenario, year), g in wide_src.groupby(["scenario", "year"], sort=False):
        row = {
            "scenario": scenario,
            "year": year,
            "pypsa_enduse_twh": g["pypsa_enduse_twh"].iloc[0],
            "peak_dso_gw": g["peak_dso_gw"].iloc[0],
            "dso_stock_gw": g["dso_stock_gw"].iloc[0],
            "dso_vs_build_pct": g["dso_vs_build_pct"].iloc[0],
            "tso_stock_vs_build_pct": g["tso_stock_vs_build_pct"].iloc[0],
        }
        for cid in [
            "A_demand_scaled",
            "B_denominator_only",
            "C_dso_capacity_all_inv",
            "D_peak_flow_all_inv",
            "E_tso_dso_split",
        ]:
            row[cid] = float(g.loc[g["case"] == cid, "system_avg_ct_per_kwh"].iloc[0])
        rows.append(row)
    print(pd.DataFrame(rows)[cols].to_string(index=False, float_format=lambda x: f"{x:,.2f}"))

    print("\nGerman DSO stock vs Build (endogenous; not the EU cap label):")
    print(
        extracts[
            [
                "scenario",
                "year",
                "dso_stock_gw",
                "peak_dso_gw",
                "dso_peak_util_pct",
                "enduse_twh",
                "eu_budget_binding",
            ]
        ].to_string(index=False, float_format=lambda x: f"{x:,.2f}")
    )


if __name__ == "__main__":
    main()
