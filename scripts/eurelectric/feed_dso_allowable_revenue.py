# SPDX-FileCopyrightText: Contributors to PyPSA-Eurelectric
#
# SPDX-License-Identifier: MIT
"""
Feed PyPSA grid-study DSO results into a copy of AllowableRevenue_MiniModelEurope.xlsx.

This script does not modify the original workbook.

Method (current):
    Build CAPEX  = original Excel CAPEX trajectory (absolute bn €/yr).
    Other CAPEX  = original Excel CAPEX × (PyPSA DSO overnight / Build overnight)
                   for that horizon. PyPSA absolute overnight is NOT injected.
    Denominator  = scenario-specific PyPSA HV→LV DSO delivery (TWh).
    RAB / WACC / OPEX / depreciation formulas stay those of the Excel.

A previous trial that replaced Excel CAPEX with PyPSA absolute overnight is
kept as ``*_TRIAL_absolute_capex.xlsx``.

Usage (repo root, pixi env):

    python scripts/eurelectric/feed_dso_allowable_revenue.py
"""

from __future__ import annotations

import argparse
import logging
import math
import re
import shutil
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pypsa

logger = logging.getLogger(__name__)

RUN_META: dict[str, dict] = {
    "grid-build-2040": {"label": "Build", "cap_fraction": 1.00},
    "grid-inaction-2040-mild85": {"label": "85%", "cap_fraction": 0.85},
    "grid-inaction-2040-mild75": {"label": "75%", "cap_fraction": 0.75},
    "grid-inaction-2040-mild65": {"label": "65%", "cap_fraction": 0.65},
    "grid-inaction-2040-moderate": {"label": "50%", "cap_fraction": 0.50},
}

# Excel Sheet1 layout (inspected, not assumed).
EXCEL_FIRST_YEAR = 2020
EXCEL_CAPEX_ROW = 17
EXCEL_YEAR_ROW = 15
EXCEL_TWH_TOTAL_ROW = 29  # F29=2020, G29=2030, H29=2040, I29=2050
EXCEL_TWH_DSO_ROW = 30
EXCEL_TARIFF_ROW = 34
# RAB block year columns: B=2020 ... L=2030 ... V=2040
YEAR_COL = {year: 2 + (year - EXCEL_FIRST_YEAR) for year in range(2020, 2059)}

ORIGINAL_TODAY_TARIFF_EUR_PER_MWH = 53.0  # Sheet1 B29, labelled "today's distribution tariff"

# Historic Excel CAPEX is kept through 2025 for every scenario.
# From 2026, Build uses the original Excel trajectory; constrained scenarios
# scale that trajectory by PyPSA DSO overnight / Build overnight.
HISTORICAL_CAPEX_THROUGH = 2025
PYPSA_2030_WINDOW = range(2026, 2031)  # 2026-2030 inclusive
PYPSA_2040_WINDOW = range(2031, 2041)  # 2031-2040 inclusive


def _weights(n: pypsa.Network) -> pd.Series:
    return n.snapshot_weightings.generators


def _weighted_sum(df: pd.DataFrame, weights: pd.Series) -> float:
    if df.empty:
        return 0.0
    return float(df.multiply(weights, axis=0).sum().sum())


def costed_dso_links(n: pypsa.Network) -> pd.Index:
    """Distribution-grid assets that carry overnight CAPEX.

    Inspected in the solved networks:

    * ``carrier == "electricity distribution grid"`` also tags data-centre
      site/demand links (``capital_cost == 0``, not extendable) and a
      reverse-direction twin of each DSO link (``capital_cost == 0``,
      ``"reversed"`` in the name). Those are not DSO investment.
    * The costed assets are one extendable link per AC node, named
      ``{node} electricity distribution grid-{year}``, with
      ``capital_cost == costs['electricity distribution grid','capital_cost']``.
    """
    dso = n.links.carrier.eq("electricity distribution grid")
    costed = n.links.capital_cost.fillna(0.0) > 0.0
    return n.links.index[dso & costed]


def dso_new_capacity_mw(n: pypsa.Network) -> pd.Series:
    """Incremental MW this horizon on costed DSO links.

    Measure: ``max(p_nom_opt - p_nom, 0)``.

    Inspected:

    * 2030: costed links are extendable, ``p_nom == 0``, ``build_year == 2030``,
      ``p_nom_min == 0``. So this is the entire modelled DSO stock. PyPSA has
      no pre-existing DSO RAB.
    * 2040: 2030 cohort is copied as non-extendable with ``p_nom = p_nom_opt``
      (see ``scripts/add_brownfield.py``). New 2040 cohort has ``p_nom == 0``.
      Incremental MW is therefore only the 2040 cohort; 2030 is not recounted.
    * ``p_nom_min`` is 0 on DSO links and is not used.
    """
    idx = costed_dso_links(n)
    if idx.empty:
        return pd.Series(dtype=float)
    dg = n.links.loc[idx]
    installed = dg["p_nom"].fillna(0.0)
    optimal = dg["p_nom_opt"].fillna(installed) if "p_nom_opt" in dg.columns else installed
    return (optimal - installed).clip(lower=0.0)


def load_dso_tech_costs(results_root: Path, run: str, year: int) -> dict:
    """Overnight investment for the distribution-grid technology.

    Source: ``resources/{prefix}/{run}/costs_{year}_processed.csv``, which is
    produced by ``scripts/process_cost_data.py`` from technology-data:

        capital_cost = (annuity(lifetime, discount_rate) + FOM/100) * investment * nyears

    Raw technology-data (archive v0.14.0): investment 667.9016 EUR/kW,
    currency_year 2015, lifetime 40, FOM 2%/yr. ``process_cost_data`` converts
    kW → MW (×1000) and does **not** inflate currency years. ``nyears`` is 1
    (snapshot weights sum to 8760 h).
    """
    resources = Path("resources") / results_root.name / run / f"costs_{year}_processed.csv"
    if not resources.exists():
        raise FileNotFoundError(
            f"Processed costs not found at {resources}. Cannot derive overnight CAPEX."
        )
    costs = pd.read_csv(resources)
    row = costs.loc[costs.technology.eq("electricity distribution grid")]
    if row.empty:
        raise ValueError(f"No 'electricity distribution grid' row in {resources}")
    r = row.iloc[0]
    return {
        "investment_eur_per_mw": float(r["investment"]),
        "fom_percent": float(r["FOM"]),
        "lifetime_years": float(r["lifetime"]),
        "discount_rate": float(r["discount rate"]),
        "capital_cost_eur_per_mw_yr": float(r["capital_cost"]),
        "currency_year": 2015,
        "raw_investment_eur_per_kw": 667.9016,
        "source": str(resources),
    }


def energy_hv_to_lv_twh(n: pypsa.Network) -> float:
    """Electricity delivered HV→LV through costed DSO links (TWh).

    Topology (``insert_electricity_distribution_grid``): costed link bus0 = AC
    node, bus1 = low-voltage node. With ``p_min_pu = 0`` after the reverse twin
    is split out, power into LV is ``-p1`` when ``p1 < 0``.

    This is energy that used the DSO asset. It excludes rooftop self-consumption
    and does not count storage charge/discharge twice. Industry electricity is
    on LV in this model, so it *is* included (unlike the Excel 80% DSO-share
    assumption). Data-centre site links are excluded (not costed).
    """
    idx = costed_dso_links(n)
    if idx.empty or n.links_t.p1.empty:
        return 0.0
    p1 = n.links_t.p1.reindex(columns=idx, fill_value=0.0)
    delivered = (-p1).clip(lower=0.0)
    return _weighted_sum(delivered, _weights(n)) / 1e6


def lv_electric_load_twh(n: pypsa.Network) -> float:
    lv_buses = n.buses.index[n.buses.carrier.eq("low voltage")]
    loads = n.loads.index[n.loads.bus.isin(lv_buses)]
    if loads.empty:
        return 0.0
    p = n.loads_t.p if not n.loads_t.p.empty else n.loads_t.p_set
    return _weighted_sum(p.reindex(columns=loads, fill_value=0.0), _weights(n)) / 1e6


def _link_elec_twh(n: pypsa.Network, mask: pd.Series, *, lv_side: str) -> float:
    """Electricity withdrawn from LV by a set of links.

    Heat pumps in these networks: bus0 = heat, bus1 = low voltage, p1 > 0 is
    withdrawal from LV. Resistive heaters / BEV chargers: bus0 = LV, p0 > 0
    is withdrawal. The previous tariff script used ``-p1`` for all bus1-LV
    links and therefore reported heat-pump electricity as 0.
    """
    names = n.links.index[mask]
    if names.empty:
        return 0.0
    w = _weights(n)
    if lv_side == "p1_positive":
        p1 = n.links_t.p1.reindex(columns=names, fill_value=0.0)
        return _weighted_sum(p1.clip(lower=0.0), w) / 1e6
    p0 = n.links_t.p0.reindex(columns=names, fill_value=0.0)
    return _weighted_sum(p0.clip(lower=0.0), w) / 1e6


def load_shedding_lv_twh(n: pypsa.Network) -> float:
    carrier_match = n.generators.carrier.str.contains("load", case=False, na=False)
    bus_lv = n.generators.bus.map(n.buses.carrier).eq("low voltage")
    gens = n.generators.index[carrier_match & bus_lv]
    if gens.empty or n.generators_t.p.empty:
        return 0.0
    p = n.generators_t.p.reindex(columns=gens, fill_value=0.0).clip(lower=0.0)
    return _weighted_sum(p, _weights(n)) / 1e6


def extract_network(n: pypsa.Network, run: str, year: int, tech: dict) -> dict:
    meta = RUN_META.get(run, {"label": run, "cap_fraction": float("nan")})
    added = dso_new_capacity_mw(n)
    added_mw = float(added.sum())
    overnight_eur = added_mw * tech["investment_eur_per_mw"]
    annualised_eur = float(
        (added * n.links.loc[added.index, "capital_cost"]).sum()
    ) if not added.empty else 0.0

    idx = costed_dso_links(n)
    dg = n.links.loc[idx] if len(idx) else n.links.iloc[0:0]
    stock_mw = float(dg["p_nom_opt"].fillna(dg["p_nom"]).sum()) if len(dg) else 0.0

    hp = _link_elec_twh(
        n, n.links.carrier.str.contains("heat pump", na=False), lv_side="p1_positive"
    )
    rh = _link_elec_twh(
        n,
        n.links.carrier.str.contains("resistive heater", na=False),
        lv_side="p0_positive",
    )
    bev = _link_elec_twh(
        n, n.links.carrier.str.contains("BEV charger", na=False), lv_side="p0_positive"
    )
    lv_load = lv_electric_load_twh(n)
    shed = load_shedding_lv_twh(n)
    lv_demand = lv_load + hp + rh + bev
    dso_twh = energy_hv_to_lv_twh(n)

    rooftop = 0.0
    rt = n.generators.index[n.generators.carrier.eq("solar rooftop")]
    if len(rt) and not n.generators_t.p.empty:
        rooftop = _weighted_sum(
            n.generators_t.p.reindex(columns=rt, fill_value=0.0).clip(lower=0.0),
            _weights(n),
        ) / 1e6

    return {
        "run": run,
        "scenario": meta["label"],
        "cap_fraction": meta["cap_fraction"],
        "year": int(year),
        "costed_dso_links": int(len(idx)),
        "new_dso_capacity_gw": added_mw / 1e3,
        "installed_dso_stock_gw": stock_mw / 1e3,
        "overnight_investment_eur_per_mw": tech["investment_eur_per_mw"],
        "overnight_capex_eur": overnight_eur,
        "overnight_capex_bn_eur": overnight_eur / 1e9,
        "annualised_new_capex_bn_eur_per_yr": annualised_eur / 1e9,
        "dso_capital_cost_eur_per_mw_yr": tech["capital_cost_eur_per_mw_yr"],
        "dso_lifetime_years": tech["lifetime_years"],
        "dso_fom_percent": tech["fom_percent"],
        "dso_discount_rate": tech["discount_rate"],
        "dso_currency_year": tech["currency_year"],
        "dso_delivered_hv_to_lv_twh": dso_twh,
        "lv_electric_load_twh": lv_load,
        "heat_pump_electricity_twh": hp,
        "resistive_heater_electricity_twh": rh,
        "bev_electricity_twh": bev,
        "electrified_lv_demand_twh": lv_demand,
        "electrified_lv_demand_served_twh": max(lv_demand - shed, 0.0),
        "lv_load_shedding_twh": shed,
        "solar_rooftop_twh": rooftop,
        "old_proxy_dso_tariff_eur_per_mwh": (
            (annualised_eur + float((dg.loc[~dg.p_nom_extendable.fillna(False), "p_nom"] * dg.loc[~dg.p_nom_extendable.fillna(False), "capital_cost"]).sum() if len(dg) else 0.0))
            / (max(lv_demand - shed, 0.0) * 1e6)
            if (lv_demand - shed) > 0
            else float("nan")
        ),
    }


def old_proxy_from_stock(n: pypsa.Network, served_lv_mwh: float) -> float:
    """Annualised cost of *all* costed DSO capacity / served LV demand.

    This is the previous PyPSA-only DSO proxy (stock, not incremental).
    """
    idx = costed_dso_links(n)
    if idx.empty or served_lv_mwh <= 0:
        return float("nan")
    dg = n.links.loc[idx]
    cap = dg["p_nom"].copy()
    ext = dg["p_nom_extendable"].fillna(False)
    cap.loc[ext] = dg.loc[ext, "p_nom_opt"].fillna(dg.loc[ext, "p_nom"])
    revenue = float((cap * dg["capital_cost"]).sum())
    return revenue / served_lv_mwh


def discover_networks(results_root: Path) -> list[tuple[str, int, Path]]:
    found = []
    for run_dir in sorted(results_root.iterdir()):
        if not run_dir.is_dir() or run_dir.name not in RUN_META:
            continue
        networks = run_dir / "networks"
        if not networks.is_dir():
            continue
        for nc in sorted(networks.glob("base_s_*_*.nc")):
            m = re.search(r"(\d{4})\.nc$", nc.name)
            if m:
                found.append((run_dir.name, int(m.group(1)), nc))
    return found


def collect_extracts(results_root: Path) -> pd.DataFrame:
    rows = []
    for run, year, path in discover_networks(results_root):
        logger.info("Reading %s", path)
        n = pypsa.Network(str(path))
        tech = load_dso_tech_costs(results_root, run, year)
        row = extract_network(n, run, year, tech)
        served = row["electrified_lv_demand_served_twh"] * 1e6
        row["old_proxy_dso_tariff_eur_per_mwh"] = old_proxy_from_stock(n, served)
        # Also report old proxy with the previous (HP-omitting) denominator
        old_denom = (
            row["lv_electric_load_twh"]
            + row["resistive_heater_electricity_twh"]
            + row["bev_electricity_twh"]
            - row["lv_load_shedding_twh"]
        ) * 1e6
        row["old_proxy_vs_incomplete_lv_demand_eur_per_mwh"] = old_proxy_from_stock(
            n, old_denom
        )
        rows.append(row)
    df = pd.DataFrame(rows)
    return df.sort_values(["year", "cap_fraction"], ascending=[True, False]).reset_index(
        drop=True
    )


def dso_overnight_ratio_vs_build(
    extracts: pd.DataFrame, scenario: str, year: int
) -> float:
    """PyPSA new DSO overnight in ``scenario`` / Build, same horizon.

    This is the only PyPSA investment signal written into Excel CAPEX.
    Build is 1.0 by definition. Absolute overnight is not used as CAPEX.
    """
    if scenario == "Build":
        return 1.0
    build = extracts[
        extracts["scenario"].eq("Build") & extracts["year"].eq(year)
    ]
    scen = extracts[
        extracts["scenario"].eq(scenario) & extracts["year"].eq(year)
    ]
    if build.empty or scen.empty:
        raise ValueError(f"Missing PyPSA extract for Build or {scenario} {year}")
    denom = float(build.iloc[0]["overnight_capex_bn_eur"])
    if denom <= 0:
        raise ValueError(f"Build overnight CAPEX is {denom} in {year}")
    return float(scen.iloc[0]["overnight_capex_bn_eur"]) / denom


def annual_capex_series(
    extracts: pd.DataFrame,
    scenario: str,
    original_capex: dict[int, float],
) -> dict[int, float]:
    """Excel row-17 CAPEX (bn EUR/yr) under the ratio-scaling method.

    * Build: no overrides (original Excel trajectory is the absolute baseline).
    * Other scenarios: original Excel CAPEX × (scenario/Build PyPSA overnight)
      for 2026–2030 (using the 2030 ratio) and 2031–2040 (using the 2040 ratio).
    * 2020–2025 and 2041+ stay at original Excel values for every scenario.
    """
    if scenario == "Build":
        return {}
    out: dict[int, float] = {}
    r2030 = dso_overnight_ratio_vs_build(extracts, scenario, 2030)
    r2040 = dso_overnight_ratio_vs_build(extracts, scenario, 2040)
    for y in PYPSA_2030_WINDOW:
        out[y] = float(original_capex[y]) * r2030
    for y in PYPSA_2040_WINDOW:
        out[y] = float(original_capex[y]) * r2040
    return out


def col_letter(idx: int) -> str:
    """1-based column index → Excel letters."""
    letters = []
    n = idx
    while n:
        n, r = divmod(n - 1, 26)
        letters.append(chr(65 + r))
    return "".join(reversed(letters))


def _mapping_rows() -> list[tuple[str, str, str, str, str]]:
    return [
        ("Excel sheet", "Cell/range", "Quantity written", "Source", "Notes"),
        (
            "PyPSA_Build",
            "H17:V17 (2026-2040)",
            "original Excel CAPEX (Build baseline)",
            "Sheet1 row 17",
            "Not replaced. PyPSA absolute overnight is not injected.",
        ),
        (
            "PyPSA_85/75/65/50pct",
            "H17:L17 (2026-2030)",
            "Excel CAPEX × (scenario 2030 overnight / Build 2030 overnight)",
            "PyPSA ratio only",
            "Typically 0.85/0.75/0.65/0.50 × 72 bn €/yr.",
        ),
        (
            "PyPSA_85/75/65/50pct",
            "M17:V17 (2031-2040)",
            "Excel CAPEX × (scenario 2040 overnight / Build 2040 overnight)",
            "PyPSA ratio only",
            "2040 cohort ratio; 2030 brownfield not recounted.",
        ),
        (
            "PyPSA_*",
            "W17:AN17 (2041+)",
            "(unchanged)",
            "Original Excel",
            "No PyPSA 2050 network; left as original 60 bn/yr for all scenarios.",
        ),
        (
            "PyPSA_*",
            "B17:G17 (2020-2025)",
            "(unchanged original historical CAPEX)",
            "Original Excel",
            "Kept so the RAB path to today is the same in every scenario.",
        ),
        (
            "PyPSA_*",
            "G30",
            "2030 DSO HV→LV delivery (TWh)",
            "weighted -min(p1,0) on costed DSO links",
            "Overwrites formula G29*$B$31 so the 80% DSO-share is NOT applied to PyPSA.",
        ),
        (
            "PyPSA_*",
            "H30",
            "2040 DSO HV→LV delivery (TWh)",
            "same",
            "Same 80% caveat.",
        ),
        (
            "PyPSA_*",
            "G35 / H35",
            "Direct DSO tariff = MAR × 1000 / TWh DSO",
            "L22 / V22 and G30 / H30",
            "This is allowable revenue / distributed electricity. Original G34/H34 still use the 2020-revenue scaling method.",
        ),
        (
            "PyPSA_*",
            "B8,B9,B11,B12,B13,B17:G17,F29,F30,I29,I30,B29",
            "(unchanged)",
            "Original Excel assumptions",
            "WACC 5.4%, lifetime 42, starting RAB 700 bn, OPEX method, 2020 volumes, today's 53 EUR/MWh.",
        ),
        (
            "PyPSA_*",
            "Rows 16,18-24",
            "(formulas unchanged)",
            "Original Excel",
            "Opening RAB, depreciation, new RAB, MAR, OPEX, ratios.",
        ),
        (
            "Original_Sheet1 / Sheet1",
            "(all)",
            "(untouched copy of the original model)",
            "Original workbook",
            "Original Excel result; do not use for PyPSA cases.",
        ),
    ]


def copy_and_fill_workbook(
    original: Path,
    dest: Path,
    extracts: pd.DataFrame,
    params: dict,
) -> None:
    shutil.copy2(original, dest)
    wb = openpyxl.load_workbook(dest)

    if "Sheet1" not in wb.sheetnames:
        raise RuntimeError(f"{original} has no Sheet1; not the expected template.")

    if "Original_Sheet1" not in wb.sheetnames:
        ws0 = wb["Sheet1"]
        copy = wb.copy_worksheet(ws0)
        copy.title = "Original_Sheet1"

    _write_inputs_sheet(wb, extracts)
    _write_mapping_sheet(wb)

    for scenario in [m["label"] for m in RUN_META.values()]:
        sub = extracts[extracts["scenario"] == scenario]
        if sub.empty:
            continue
        title = f"PyPSA_{scenario.replace('%', 'pct')}"
        if title in wb.sheetnames:
            del wb[title]
        ws = wb.copy_worksheet(wb["Original_Sheet1"])
        ws.title = title
        _fill_scenario_sheet(ws, extracts, sub, scenario, params["original_capex"])

    wb.save(dest)
    logger.info("Wrote %s", dest)


def _write_inputs_sheet(wb: openpyxl.Workbook, extracts: pd.DataFrame) -> None:
    if "Inputs_PyPSA" in wb.sheetnames:
        del wb["Inputs_PyPSA"]
    ws = wb.create_sheet("Inputs_PyPSA", 0)
    ws["A1"] = "Raw quantities extracted from solved PyPSA networks"
    ws["A2"] = (
        "DSO assets: costed links with carrier 'electricity distribution grid' "
        "(capital_cost > 0). Overnight CAPEX is used only as a ratio vs Build; "
        "it is not written into Excel as an absolute bn EUR figure."
    )
    # write dataframe
    cols = list(extracts.columns)
    for j, c in enumerate(cols, 1):
        ws.cell(4, j, c)
    for i, row in enumerate(extracts.itertuples(index=False), 5):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, None if (isinstance(val, float) and math.isnan(val)) else val)


def _write_mapping_sheet(wb: openpyxl.Workbook) -> None:
    if "Mapping" in wb.sheetnames:
        del wb["Mapping"]
    ws = wb.create_sheet("Mapping", 1)
    for i, row in enumerate(_mapping_rows(), 1):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)


def _fill_scenario_sheet(
    ws,
    extracts: pd.DataFrame,
    sub: pd.DataFrame,
    scenario: str,
    original_capex: dict[int, float],
) -> None:
    by_year = sub.set_index("year")
    capex = annual_capex_series(extracts, scenario, original_capex)
    for year, value in capex.items():
        col = YEAR_COL[year]
        ws.cell(EXCEL_CAPEX_ROW, col, value)

    # DSO delivered TWh into the TWh DSO row (G30=2030, H30=2040).
    if 2030 in by_year.index:
        ws.cell(EXCEL_TWH_DSO_ROW, 7, float(by_year.loc[2030, "dso_delivered_hv_to_lv_twh"]))
        ws.cell(EXCEL_TWH_TOTAL_ROW, 7, float(by_year.loc[2030, "dso_delivered_hv_to_lv_twh"]))
    if 2040 in by_year.index:
        ws.cell(EXCEL_TWH_DSO_ROW, 8, float(by_year.loc[2040, "dso_delivered_hv_to_lv_twh"]))
        ws.cell(EXCEL_TWH_TOTAL_ROW, 8, float(by_year.loc[2040, "dso_delivered_hv_to_lv_twh"]))

    r2030 = dso_overnight_ratio_vs_build(extracts, scenario, 2030)
    r2040 = dso_overnight_ratio_vs_build(extracts, scenario, 2040)
    ws["E35"] = "Direct DSO tariff MAR/TWh (EUR/MWh)"
    ws["F35"] = "=B22*1000/F30"
    ws["G35"] = "=L22*1000/G30"
    ws["H35"] = "=V22*1000/H30"
    ws["A50"] = (
        f"PyPSA scenario {scenario}: Build CAPEX = original Excel trajectory. "
        f"This sheet CAPEX 2026-2030 = Excel × {r2030:.4f}; "
        f"2031-2040 = Excel × {r2040:.4f} "
        "(PyPSA DSO overnight / Build). Absolute PyPSA overnight is not injected. "
        "G30/H30 are scenario HV→LV TWh (no 80% factor). "
        "2020-2025 and 2041+ CAPEX unchanged. RAB/WACC/OPEX formulas unchanged."
    )


def excel_params_from_original(original: Path) -> dict:
    wb = openpyxl.load_workbook(original, data_only=False)
    ws = wb["Sheet1"]
    capex = {}
    for year, col in YEAR_COL.items():
        capex[year] = ws.cell(EXCEL_CAPEX_ROW, col).value
    return {
        "wacc": float(ws["B8"].value),
        "lifetime": float(ws["B9"].value),
        "capex_share": float(ws["B11"].value),
        "opex_growth": float(ws["B12"].value),
        "start_rab": float(ws["B13"].value),
        "today_tariff": float(ws["B29"].value),
        "twh_2020_total": float(ws["F29"].value),
        "twh_2030_total_orig": float(ws["G29"].value),
        "twh_2040_total_orig": float(ws["H29"].value),
        "dso_share": float(ws["B31"].value),
        "est_revenue_2020_formula_inputs": {
            "consumption_twh": float(ws["B30"].value),
        },
        "original_capex": capex,
        "freeze_depreciation_after": 2030,  # Sheet1 M18=L18
    }


def simulate_sheet1(
    params: dict,
    capex_override: dict[int, float] | None,
    twh_dso: dict[int, float],
) -> pd.DataFrame:
    """Replicate Sheet1 identities in Python."""
    wacc = params["wacc"]
    dep_rate = 1.0 / params["lifetime"]
    capex_share = params["capex_share"]
    g_opex = params["opex_growth"]
    years = list(range(2020, 2059))
    original_capex = params["original_capex"]

    opening = {}
    capex = {}
    dep = {}
    new_rab = {}
    mar_capex = {}
    opex = {}
    mar = {}
    dep_2030 = None

    for i, year in enumerate(years):
        if year == 2020:
            opening[year] = params["start_rab"]
        else:
            opening[year] = new_rab[year - 1]

        raw = original_capex[year]
        if capex_override is not None and year in capex_override:
            capex[year] = capex_override[year]
        else:
            capex[year] = float(raw)

        if year <= params["freeze_depreciation_after"]:
            dep[year] = opening[year] * dep_rate
            if year == params["freeze_depreciation_after"]:
                dep_2030 = dep[year]
        else:
            dep[year] = dep_2030

        new_rab[year] = opening[year] + capex[year] - dep[year]
        mar_capex[year] = (opening[year] + capex[year]) * wacc + dep[year]
        if year == 2020:
            opex[year] = mar_capex[year] * (1.0 / capex_share - 1.0)
            opex_2020 = opex[year]
        else:
            opex[year] = opex_2020 * (1.0 + g_opex) ** (year - 2020)
        mar[year] = mar_capex[year] + opex[year]

    est_rev_2020 = (
        params["est_revenue_2020_formula_inputs"]["consumption_twh"]
        * params["dso_share"]
        * 1e6
        * params["today_tariff"]
        * 1e-9
    )

    rows = []
    for year in years:
        twh = twh_dso.get(year)
        scaled_rev = est_rev_2020 * (mar[year] / mar[2020])
        direct = mar[year] * 1000.0 / twh if twh else float("nan")
        scaled = scaled_rev * 1000.0 / twh if twh else float("nan")
        rows.append(
            {
                "year": year,
                "opening_rab_bn": opening[year],
                "capex_bn": capex[year],
                "depreciation_bn": dep[year],
                "new_rab_bn": new_rab[year],
                "mar_capex_bn": mar_capex[year],
                "opex_bn": opex[year],
                "allowable_revenue_bn": mar[year],
                "twh_dso": twh,
                "tariff_direct_eur_per_mwh": direct,
                "tariff_scaled_from_2020_eur_per_mwh": scaled,
                "mar_index_2020": mar[year] / mar[2020],
            }
        )
    return pd.DataFrame(rows)


def interpolate_twh(twh_at: dict[int, float], years: range) -> dict[int, float]:
    knots = sorted(twh_at)
    out = {}
    for y in years:
        if y in twh_at:
            out[y] = twh_at[y]
            continue
        below = [k for k in knots if k < y]
        above = [k for k in knots if k > y]
        if below and above:
            y0, y1 = below[-1], above[0]
            out[y] = twh_at[y0] + (twh_at[y1] - twh_at[y0]) * (y - y0) / (y1 - y0)
        elif below:
            out[y] = twh_at[below[-1]]
        elif above:
            out[y] = twh_at[above[0]]
    return out


def build_comparison(
    extracts: pd.DataFrame, params: dict
) -> tuple[pd.DataFrame, pd.DataFrame]:
    orig_twh = {
        2020: params["twh_2020_total"] * params["dso_share"],
        2030: params["twh_2030_total_orig"] * params["dso_share"],
        2040: params["twh_2040_total_orig"] * params["dso_share"],
    }
    orig_sim = simulate_sheet1(params, capex_override=None, twh_dso=interpolate_twh(orig_twh, range(2020, 2051)))
    orig_sim.insert(0, "scenario", "Original Excel")

    frames = [orig_sim]
    snapshot_rows = []

    for year in (2030, 2040):
        orow = orig_sim[orig_sim["year"] == year].iloc[0]
        snapshot_rows.append(
            {
                "scenario": "Original Excel",
                "year": year,
                "new_dso_investment_bn_eur_per_yr": orow["capex_bn"],
                "dso_rab_bn_eur": orow["new_rab_bn"],
                "allowable_dso_revenue_bn_eur_per_yr": orow["allowable_revenue_bn"],
                "distributed_electricity_twh": orow["twh_dso"],
                "dso_tariff_eur_per_mwh": orow["tariff_direct_eur_per_mwh"],
                "dso_tariff_scaled_excel_method_eur_per_mwh": orow[
                    "tariff_scaled_from_2020_eur_per_mwh"
                ],
                "original_excel_tariff_eur_per_mwh": orow[
                    "tariff_scaled_from_2020_eur_per_mwh"
                ],
                "old_pypsa_proxy_eur_per_mwh": float("nan"),
                "old_pypsa_proxy_incl_hp_eur_per_mwh": float("nan"),
                "overnight_capex_this_horizon_bn": float("nan"),
                "new_dso_capacity_gw": float("nan"),
                "pypsa_capex_ratio_vs_build": 1.0,
                "today_tariff_eur_per_mwh": params["today_tariff"],
            }
        )

    for scenario in [m["label"] for m in RUN_META.values()]:
        sub = extracts[extracts["scenario"] == scenario]
        if sub.empty:
            continue
        by_year = sub.set_index("year")
        twh_knots = {
            2020: orig_twh[2020],
        }
        if 2030 in by_year.index:
            twh_knots[2030] = float(by_year.loc[2030, "dso_delivered_hv_to_lv_twh"])
        if 2040 in by_year.index:
            twh_knots[2040] = float(by_year.loc[2040, "dso_delivered_hv_to_lv_twh"])
        twh = interpolate_twh(twh_knots, range(2020, 2051))
        capex = annual_capex_series(extracts, scenario, params["original_capex"])
        sim = simulate_sheet1(params, capex_override=capex, twh_dso=twh)
        sim.insert(0, "scenario", scenario)
        frames.append(sim)

        for year in (2030, 2040):
            if year not in by_year.index:
                continue
            srow = sim[sim["year"] == year].iloc[0]
            orow = orig_sim[orig_sim["year"] == year].iloc[0]
            snapshot_rows.append(
                {
                    "scenario": scenario,
                    "year": year,
                    "new_dso_investment_bn_eur_per_yr": srow["capex_bn"],
                    "dso_rab_bn_eur": srow["new_rab_bn"],
                    "allowable_dso_revenue_bn_eur_per_yr": srow["allowable_revenue_bn"],
                    "distributed_electricity_twh": srow["twh_dso"],
                    "dso_tariff_eur_per_mwh": srow["tariff_direct_eur_per_mwh"],
                    "dso_tariff_scaled_excel_method_eur_per_mwh": srow[
                        "tariff_scaled_from_2020_eur_per_mwh"
                    ],
                    "original_excel_tariff_eur_per_mwh": orow[
                        "tariff_scaled_from_2020_eur_per_mwh"
                    ],
                    "old_pypsa_proxy_eur_per_mwh": float(
                        by_year.loc[year, "old_proxy_vs_incomplete_lv_demand_eur_per_mwh"]
                    ),
                    "old_pypsa_proxy_incl_hp_eur_per_mwh": float(
                        by_year.loc[year, "old_proxy_dso_tariff_eur_per_mwh"]
                    ),
                    "overnight_capex_this_horizon_bn": float(
                        by_year.loc[year, "overnight_capex_bn_eur"]
                    ),
                    "new_dso_capacity_gw": float(by_year.loc[year, "new_dso_capacity_gw"]),
                    "pypsa_capex_ratio_vs_build": dso_overnight_ratio_vs_build(
                        extracts, scenario, year
                    ),
                    "today_tariff_eur_per_mwh": params["today_tariff"],
                }
            )

    annual = pd.concat(frames, ignore_index=True)
    snap = pd.DataFrame(snapshot_rows)

    # vs 2025 / vs Build
    def _sim(scenario, year):
        m = annual[(annual["scenario"] == scenario) & (annual["year"] == year)]
        return m.iloc[0] if not m.empty else None

    extras = []
    for _, r in snap.iterrows():
        s2025 = _sim(r["scenario"], 2025)
        build = _sim("Build", int(r["year"]))
        orig = _sim("Original Excel", int(r["year"]))
        tariff = r["dso_tariff_eur_per_mwh"]
        extras.append(
            {
                "tariff_change_vs_today53_eur_per_mwh": tariff - params["today_tariff"],
                "tariff_change_vs_today53_pct": 100.0
                * (tariff / params["today_tariff"] - 1.0),
                "tariff_change_vs_2025_eur_per_mwh": (
                    tariff - s2025["tariff_direct_eur_per_mwh"] if s2025 is not None else float("nan")
                ),
                "tariff_change_vs_2025_pct": (
                    100.0 * (tariff / s2025["tariff_direct_eur_per_mwh"] - 1.0)
                    if s2025 is not None
                    else float("nan")
                ),
                "tariff_vs_build_eur_per_mwh": (
                    0.0
                    if r["scenario"] in {"Build", "Original Excel"} or build is None
                    else tariff - build["tariff_direct_eur_per_mwh"]
                ),
                "tariff_vs_original_excel_eur_per_mwh": (
                    0.0
                    if r["scenario"] == "Original Excel" or orig is None
                    else tariff - orig["tariff_direct_eur_per_mwh"]
                ),
                "investment_vs_build_bn_eur_per_yr": (
                    0.0
                    if r["scenario"] in {"Build", "Original Excel"} or build is None
                    else r["new_dso_investment_bn_eur_per_yr"] - build["capex_bn"]
                ),
                "distributed_twh_vs_build": (
                    0.0
                    if r["scenario"] in {"Build", "Original Excel"} or build is None
                    else r["distributed_electricity_twh"] - build["twh_dso"]
                ),
                "rab_vs_build_bn": (
                    0.0
                    if r["scenario"] in {"Build", "Original Excel"} or build is None
                    else r["dso_rab_bn_eur"] - build["new_rab_bn"]
                ),
            }
        )
    snap = pd.concat([snap.reset_index(drop=True), pd.DataFrame(extras)], axis=1)
    return snap, annual


METHODOLOGY = """# PyPSA → DSO allowable-revenue model

## Excel model (inspected, not modified)

File: `AllowableRevenue_MiniModelEurope.xlsx`, sheet **Sheet1**
(Sheet1 (2) is a depreciation variant; Sheet1 is the working model).

| Item | Cell | Kind | Value / formula | Units |
|---|---|---|---|---|
| WACC | B8 | assumption | 0.054 | fraction |
| Average asset lifetime | B9 | assumption | 42 | years |
| Depreciation rate | B10 | formula | `=1/B9` | 1/year |
| CAPEX share of allowable revenue | B11 | assumption | 0.5 | fraction |
| OPEX growth | B12 | assumption | 0.01 | 1/year |
| Starting RAB Europe | B13 | assumption | 700 | billion EUR |
| Opening RAB | row 16 | formula | 2020=`B13`; later=`previous New RAB` | billion EUR |
| CAPEX addition | row 17 | assumption | 29…60 then 72 (2026-2040) then 60 | billion EUR/yr |
| Depreciation | row 18 | formula | `Opening RAB × dep rate` through 2030; **frozen at 2030 from 2031** (`M18=L18`) | billion EUR/yr |
| New RAB | row 19 | formula | `Opening + CAPEX − Depreciation` | billion EUR |
| Max allowable revenue (CAPEX part) | row 20 | formula | `(Opening+CAPEX)×WACC + Depreciation` | billion EUR/yr |
| OPEX | row 21 | formula | 2020: `MAR_capex×(1/share−1)`; later: 2020 OPEX grown at 1%/yr | billion EUR/yr |
| Max allowable revenue | row 22 | formula | CAPEX-part + OPEX | billion EUR/yr |
| Today's DSO tariff | B29 | assumption | 53 | EUR/MWh |
| 2020 electricity consumption | B30 / F29 | assumption | 2590 / 2589.57 | TWh |
| DSO share of consumption | B31 | assumption | 0.8 | fraction |
| TWh DSO | row 30 | formula | `TWh × 0.8` | TWh |
| Estimated 2020 DSO revenue | B32 | formula | `2590×0.8×53 / 1000` = 109.816 | billion EUR |
| Distribution charges | F34:I34 | formula | scaled 2020 revenue × MAR index / TWh DSO | EUR/MWh |

No price year is stated in the workbook. Cached original tariffs: 2020 ≈ 53.0, 2030 ≈ 48.0, 2040 ≈ 52.0 EUR/MWh.

## PyPSA DSO investment (inspected)

Assets used: links with `carrier == "electricity distribution grid"` **and**
`capital_cost > 0`.

Excluded (same carrier, not DSO investment):

- Data-centre site/demand links (`capital_cost = 0`, not extendable).
- Reverse-direction twins (`"reversed"` in the name, `capital_cost = 0`).

New capacity this horizon:

    ΔP_MW = max(p_nom_opt − p_nom, 0)   on costed links only

| Horizon | What the networks show |
|---|---|
| 2030 | Costed links are new (`p_nom=0`, `p_nom_min=0`, `build_year=2030`, extendable). PyPSA has **no existing DSO RAB**. ΔP is the entire modelled 2030 DSO overlay. |
| 2040 | 2030 cohort is brownfield (`p_nom = 2030 p_nom_opt`, not extendable). 2040 cohort is new. ΔP is **only the 2040 additions**. |

Overnight CAPEX (cash, used **only as a ratio vs Build**):

    overnight EUR = ΔP_MW × 667 901.6 EUR/MW
    ratio_s,y = overnight_{s,y} / overnight_{Build,y}

667 901.6 EUR/MW comes from processed technology-data (`investment` after kW→MW).
Raw value: 667.9016 EUR/kW, **currency_year 2015**, lifetime 40 years, FOM 2%/year,
discount rate 7%. `capital_cost` in PyPSA **includes FOM**. Absolute overnight is
**not** written into the Excel CAPEX row.

### CAPEX written into Excel (ratio method)

A previous trial injected PyPSA absolute overnight into row 17. That workbook is
archived as `AllowableRevenue_MiniModelEurope_PyPSA_TRIAL_absolute_capex.xlsx`.

Current method:

- **Original Excel / Build:** the Excel CAPEX trajectory is the absolute baseline
  (29→60 then 72 bn €/yr in 2026–2040, then 60). Build does not replace those
  numbers with PyPSA euros.
- **85% / 75% / 65% / 50%:**
  - 2020–2025: original Excel CAPEX (same for every scenario)
  - 2026–2030: Excel CAPEX × ratio_2030
  - 2031–2040: Excel CAPEX × ratio_2040
  - 2041+: original Excel 60 bn/yr
- **TWh DSO (G30/H30):** scenario-specific PyPSA HV→LV delivery for *all*
  PyPSA sheets, including Build. Original Excel keeps its own 80%-of-consumption volumes.

This avoids mixing PyPSA's thin HV–LV overlay (EUR2015) with the Excel's 700 bn
European DSO RAB. PyPSA only answers *how much less* is built under inaction.

## Electricity denominator (inspected)

`insert_electricity_distribution_grid` puts regular electricity, industry
electricity, agriculture, BEV chargers, heat pumps and resistive heaters on
**low-voltage** buses. The costed DSO link is the HV (AC) → LV connection.

Chosen billing volume for the Excel (G30/H30):

    DSO delivered TWh = snapshot-weighted sum of max(−p1, 0) on costed DSO links

That is energy that crossed the DSO asset (HV→LV). It is the quantity that
matches "electricity distributed through the DSO grid".

It is **not** total generation. It does **not** apply Excel's 80% DSO share
(industry is already behind the DSO link in PyPSA). Rooftop PV serving local
load does not cross the link and is excluded.

### Sanity check vs the old 3,175 / 4,082 TWh

Those figures were `LV loads + resistive heaters + BEV chargers − LV shedding`.
Heat-pump electricity was **omitted** (the old script took `−p1` on bus1=LV, but
heat pumps have `p1 > 0`). Independently:

| | 2030 Build | 2040 Build |
|---|---:|---:|
| Old incomplete LV demand | ~3,175 TWh | ~4,082 TWh |
| Heat-pump electricity (p1>0) | ~631 TWh | computed in CSV |
| DSO HV→LV (this study) | ~3,082 TWh | ~3,927 TWh |
| Rooftop generation | ~722 TWh | ~1,026 TWh |

DSO flow + rooftop ≈ LV demand including heat pumps. That closes the energy
balance and is why the DSO volume is below 3,175+HP.

## What is written into Excel

See sheet `Mapping` in the copied workbook. Formulas for RAB, depreciation,
OPEX and allowable revenue are **not** overwritten. Build CAPEX is the original
Excel trajectory. Constrained sheets only scale 2026–2040 CAPEX by the PyPSA
overnight ratio. 2030/2040 TWh DSO are PyPSA HV→LV volumes.

Direct tariff added in G35/H35:

    EUR/MWh = allowable revenue (bn EUR) × 1000 / TWh DSO

## Why the old PyPSA proxy is smaller

Old proxy ≈ annualised *modelled* DSO stock / LV kWh ≈ 9.7–11 EUR/MWh.
It only recovers the HV–LV overlay that PyPSA represents (~31–45 bn €/yr
annualised), not the 700 bn existing RAB, not Excel OPEX, not 5.4% WACC on
the full RAB.

The regulatory tariff recovers `RAB×WACC + depreciation + OPEX` on the Excel
RAB, with inaction represented as a *fraction* of the Excel CAPEX path.

## Needs confirmation

1. **2041+ CAPEX** left at original 60 bn/yr for every scenario.
2. **Sheet1 depreciation freeze** after 2030 was kept.
3. **Excel 80% DSO share** is *not* applied to PyPSA volumes (industry is on LV).
4. Geography: PyPSA is EU27+UK+NO+CH; Excel says "Europe".
5. Build vs Original Excel now differ only in the TWh denominator (PyPSA HV→LV
   vs Excel consumption×0.8), not in CAPEX.
"""


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    p.add_argument(
        "--results-root",
        type=Path,
        default=Path("results/2026-grid-study"),
    )
    p.add_argument(
        "--original-xlsx",
        type=Path,
        default=Path(
            "results/2026-grid-study/distribution_tariffs/AllowableRevenue_MiniModelEurope.xlsx"
        ),
    )
    p.add_argument(
        "--output-dir",
        type=Path,
        default=Path("results/2026-grid-study/distribution_tariffs/dso_allowable_revenue"),
    )
    p.add_argument(
        "--extract-csv",
        type=Path,
        default=None,
        help="Reuse a previously written pypsa_dso_extract.csv instead of re-reading networks.",
    )
    return p.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    if args.extract_csv is not None:
        extracts = pd.read_csv(args.extract_csv)
        logger.info("Reused PyPSA extract from %s", args.extract_csv)
    else:
        extracts = collect_extracts(args.results_root)
    extracts_path = args.output_dir / "pypsa_dso_extract.csv"
    extracts.to_csv(extracts_path, index=False, float_format="%.6f")
    logger.info("Wrote %s", extracts_path)

    params = excel_params_from_original(args.original_xlsx)
    snap, annual = build_comparison(extracts, params)

    snap_path = args.output_dir / "dso_tariff_comparison.csv"
    snap.to_csv(snap_path, index=False, float_format="%.4f")
    logger.info("Wrote %s", snap_path)

    annual_path = args.output_dir / "allowable_revenue_annual.csv"
    annual.to_csv(annual_path, index=False, float_format="%.4f")
    logger.info("Wrote %s", annual_path)

    xlsx_path = args.output_dir / "AllowableRevenue_MiniModelEurope_PyPSA.xlsx"
    trial_path = (
        args.output_dir / "AllowableRevenue_MiniModelEurope_PyPSA_TRIAL_absolute_capex.xlsx"
    )
    if xlsx_path.exists() and not trial_path.exists():
        shutil.move(str(xlsx_path), str(trial_path))
        note = args.output_dir / "TRIAL_absolute_capex_README.txt"
        note.write_text(
            "TRIAL: this workbook injected PyPSA absolute DSO overnight CAPEX "
            "into Excel row 17. It is kept for reference only. The live workbook "
            "AllowableRevenue_MiniModelEurope_PyPSA.xlsx uses the original Excel "
            "CAPEX trajectory as the Build baseline and scales it by PyPSA "
            "scenario/Build DSO overnight ratios.\n"
        )
        logger.info("Archived previous absolute-CAPEX workbook as %s", trial_path)

    copy_and_fill_workbook(args.original_xlsx, xlsx_path, extracts, params)

    md_path = args.output_dir / "methodology.md"
    md_path.write_text(METHODOLOGY)
    logger.info("Wrote %s", md_path)

    mapping_path = args.output_dir / "excel_cell_mapping.csv"
    mapping_rows = _mapping_rows()
    pd.DataFrame(mapping_rows[1:], columns=list(mapping_rows[0])).to_csv(
        mapping_path, index=False
    )
    logger.info("Wrote %s", mapping_path)

    cols = [
        "scenario",
        "year",
        "new_dso_investment_bn_eur_per_yr",
        "pypsa_capex_ratio_vs_build",
        "dso_rab_bn_eur",
        "allowable_dso_revenue_bn_eur_per_yr",
        "distributed_electricity_twh",
        "dso_tariff_eur_per_mwh",
        "dso_tariff_scaled_excel_method_eur_per_mwh",
        "old_pypsa_proxy_eur_per_mwh",
        "tariff_vs_build_eur_per_mwh",
        "tariff_vs_original_excel_eur_per_mwh",
    ]
    print(snap[cols].to_string(index=False, float_format=lambda x: f"{x:,.2f}"))

    print("\nRaw PyPSA extract (Build):")
    print(
        extracts.loc[
            extracts.scenario.eq("Build"),
            [
                "year",
                "new_dso_capacity_gw",
                "overnight_capex_bn_eur",
                "dso_delivered_hv_to_lv_twh",
                "electrified_lv_demand_twh",
                "heat_pump_electricity_twh",
            ],
        ].to_string(index=False, float_format=lambda x: f"{x:,.2f}")
    )


if __name__ == "__main__":
    main()
